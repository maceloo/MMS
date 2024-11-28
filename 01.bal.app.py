import tkinter as tk
from tkinter import ttk
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os

class AfericaoBalancaApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Registro de Aferição de Balanças")

        # Definindo cores
        self.cor_fundo = "#E8F5E9"
        self.cor_titulo = "#388E3C"
        self.cor_label_frame = "#C8E6C9"
        self.cor_botao = "#2E7D32"

        # Frame principal
        self.frame_principal = tk.Frame(master, padx=10, pady=10, bg=self.cor_fundo)
        self.frame_principal.pack(fill=tk.BOTH, expand=True)

        # Título
        tk.Label(self.frame_principal, text="Registro de Aferição", font=("Helvetica", 16, "bold"), bg=self.cor_fundo, fg=self.cor_titulo).grid(row=0, columnspan=2, pady=10)

        # Frame de Data e Balança
        self.frame_dados = tk.LabelFrame(self.frame_principal, text="Dados da Aferição", padx=10, pady=10, bg=self.cor_label_frame)
        self.frame_dados.grid(row=1, column=0, columnspan=2, pady=10)

        # Data atual
        self.data_atual = tk.StringVar()
        self.data_atual.set(datetime.now().strftime("%d/%m/%Y"))
        tk.Label(self.frame_dados, text="Data:", bg=self.cor_label_frame).grid(row=0, column=0)
        tk.Entry(self.frame_dados, textvariable=self.data_atual, state='readonly', width=20).grid(row=0, column=1)

        # Escolha da balança
        self.balanca_selecionada = tk.StringVar()
        tk.Label(self.frame_dados, text="Escolha a balança:", bg=self.cor_label_frame).grid(row=1, column=0)
        self.balanca_combobox = ttk.Combobox(self.frame_dados, textvariable=self.balanca_selecionada, width=18)
        self.balanca_combobox['values'] = (
            "Balança 01", "Balança 02", "Balança 03", "Balança 04", 
            "Balança 05", "Balança 06", "Balança 07", "Balança 08")
        self.balanca_combobox.grid(row=1, column=1, pady=5)
        
        # Adicionar evento para limpar campos ao mudar a balança selecionada
        self.balanca_combobox.bind("<<ComboboxSelected>>", self.limpar_campos)

        # Frame de Pesos
        self.frame_pesos = tk.LabelFrame(self.frame_principal, text="Pesos Registrados", padx=10, pady=10, bg=self.cor_label_frame)
        self.frame_pesos.grid(row=2, column=0, columnspan=2, pady=10)

        # Pesos
        self.peso1 = tk.StringVar()
        self.peso2 = tk.StringVar()
        self.peso3 = tk.StringVar()
        
        tk.Label(self.frame_pesos, text="Peso 1 (g):", bg=self.cor_label_frame).grid(row=0, column=0)
        tk.Entry(self.frame_pesos, textvariable=self.peso1, width=20).grid(row=0, column=1)
        
        tk.Label(self.frame_pesos, text="Peso 2 (g):", bg=self.cor_label_frame).grid(row=1, column=0)
        tk.Entry(self.frame_pesos, textvariable=self.peso2, width=20).grid(row=1, column=1)
        
        tk.Label(self.frame_pesos, text="Peso 3 (g):", bg=self.cor_label_frame).grid(row=2, column=0)
        tk.Entry(self.frame_pesos, textvariable=self.peso3, width=20).grid(row=2, column=1)
        
        # Frame de Resultados
        self.frame_resultados = tk.LabelFrame(self.frame_principal, text="Resultados", padx=10, pady=10, bg=self.cor_label_frame)
        self.frame_resultados.grid(row=3, column=0, columnspan=2, pady=10)

        # Botão para calcular a média e o limite de erro
        self.botao_calcular = tk.Button(self.frame_resultados, text="Calcular Média e Variação de Erro", command=self.calcular_media, bg=self.cor_botao, fg="white")
        self.botao_calcular.grid(row=0, columnspan=2, pady=5)
        
        # Resultado da média
        self.resultado_media = tk.StringVar()
        tk.Label(self.frame_resultados, text="Média dos Pesos (g):", bg=self.cor_label_frame).grid(row=1, column=0)
        tk.Entry(self.frame_resultados, textvariable=self.resultado_media, state='readonly', width=20).grid(row=1, column=1)
        
        # Limite de erro
        self.limite_erro = tk.StringVar()
        tk.Label(self.frame_resultados, text="Variação de Peso (g):", bg=self.cor_label_frame).grid(row=2, column=0)
        tk.Entry(self.frame_resultados, textvariable=self.limite_erro, state='readonly', width=20).grid(row=2, column=1)
        
        # Botão para salvar em Excel
        self.botao_salvar = tk.Button(self.frame_principal, text="Salvar", command=self.salvar_excel, bg=self.cor_botao, fg="white")
        self.botao_salvar.grid(row=4, columnspan=2, pady=10)

    def calcular_media(self):
        try:
            peso1 = float(self.peso1.get())
            peso2 = float(self.peso2.get())
            peso3 = float(self.peso3.get())
            # Define o peso de referência com base na balança selecionada
            if self.balanca_selecionada.get() == "Balança 04":
                peso_referencia = 1000.0  # 1kg
            else:
                peso_referencia = 200.0  # 200g

            media = (peso1 + peso2 + peso3) / 3
            self.resultado_media.set(f"{media:.2f} g")

            # Calculando o limite de erro
            limite_erro = media - peso_referencia
            self.limite_erro.set(f"{limite_erro:.2f} g")
        except Exception as e:
            self.resultado_media.set(f"Erro: {e}")

    def salvar_excel(self):
        try:
            # Definir o nome do arquivo Excel
            arquivo_excel = "registro_afericoes.xlsx"

            # Verificar se o arquivo já existe
            if os.path.exists(arquivo_excel):
                workbook = load_workbook(arquivo_excel)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove a planilha padrão

            # Criar uma nova planilha para a balança selecionada
            sheet = workbook.create_sheet(self.balanca_selecionada.get())
            sheet.title = self.balanca_selecionada.get()

            # Adicionar os dados ao Excel
            dados = [
                ["Data", self.data_atual.get()],
                ["Balança", self.balanca_selecionada.get()],
                ["Peso 01 (g)", self.peso1.get()],
                ["Peso 02 (g)", self.peso2.get()],
                ["Peso 03 (g)", self.peso3.get()],
                ["Média dos Pesos (g)", self.resultado_media.get()],
                ["Variação de Peso (g)", self.limite_erro.get()]
            ]

            for row in dados:
                sheet.append(row)
            
            # Estilizar as células
            italic_font = Font(italic=True)
            center_alignment = Alignment(horizontal="center")
            
            for row in sheet.iter_rows(min_row=1, max_row=len(dados), min_col=1, max_col=2):
                for cell in row:
                    cell.font = italic_font
                    cell.alignment = center_alignment

            # Salvar o arquivo
            workbook.save(arquivo_excel)
            print("Arquivo Excel salvo com sucesso!")

            # Limpar os campos de entrada
            self.limpar_campos()
        except Exception as e:
            print(f"Erro ao salvar o arquivo Excel: {e}")

    def limpar_campos(self, event=None):
        self.peso1.set("")
        self.peso2.set("")
        self.peso3.set("")

if __name__ == "__main__":
    root = tk.Tk()
    app = AfericaoBalancaApp(root)
    root.mainloop()

        
        